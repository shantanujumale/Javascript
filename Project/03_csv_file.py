from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "form_data.xlsx"

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json  # Expect JSON from frontend
    name = data.get("name")
    age = data.get("age")
    email = data.get("email")
    profile = data.get("profile")

    # If file exists, load it; otherwise create new workbook
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Email", "Profile Pic"])  # header row

    # Append new row
    ws.append([name, age, email, profile])
    wb.save(FILE_NAME)

    return jsonify({"message": "Data saved successfully!"})

if __name__ == "__main__":
    app.run(debug=True)
